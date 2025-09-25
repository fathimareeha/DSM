from django.shortcuts import render
from rest_framework import generics
from .serializer import UserSerializer,SchoolSerializer,CollegeSerializer,InstitutionAdminLoginSerializer,Subscription_packageSerializer,PaymentSerializer,AllInstitutionSerializer,NotificationSerializer,CourseSerializer,DepartmentSerializer,SemesterSerializer,SubjectSerializer,UniversitySerializer,LandingPageContentSerializer,UniversityNestedSerializer,CouponSerializer
from .models import UserProfile,School,College,Institution,SubscriptionPackage,Payment,Notification,StaffRole,Course,Department,Semester,Subject,University,LandingPageContent,Coupon,PendingPayment,PaymentHistory
from rest_framework import authentication,permissions,serializers
from rest_framework.views import APIView 
from rest_framework.authtoken.models import Token
from rest_framework.response import Response
from .permission import IsInstitutionAdmin,IsSuperadminOrStaff,IsSuperadminOrReadOnlyForStaff,IsSuperAdminOrInstitutionManager,IsSuperAdminOrPackageManager,IsSuperAdminOrAcademicManager
import razorpay
from datetime import timedelta
from django.utils import timezone
from django.core.mail import send_mail
from django_rest_passwordreset.signals import reset_password_token_created
from django.dispatch import receiver
from rest_framework import status
  # or your custom user model
from django.contrib.auth.hashers import make_password
from django.shortcuts import get_object_or_404
from django.contrib.auth import get_user_model, authenticate
from .serializer import AdminLoginSerializer
from django.db.models import Sum, F
from rest_framework.decorators import api_view
from datetime import date
from decimal import Decimal
from django.contrib.auth.tokens import default_token_generator
from superadmin_app.constants.pincodes import PINCODE_CHOICES
User = get_user_model()


SECRET_KEY='scF0IXTa15CIYsinEfjDaa5D' # put the downloaded secretkey and key_id here 
KEY_ID='rzp_test_1GToET5GDSxcq2'

# Create your views here.


class CreateUserView(generics.ListCreateAPIView):
    
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[IsSuperAdminOrInstitutionManager]
    serializer_class=UserSerializer
    queryset=UserProfile.objects.all()
    
    def perform_create(self, serializer):
        serializer.save(role='institution_admin')
    
    def get_queryset(self):
        return UserProfile.objects.filter(role="institution_admin")
    

class DeleteInstitutionAndAdmin(generics.DestroyAPIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperAdminOrInstitutionManager]

    def delete(self, request, pk):
        try:
            institution = Institution.objects.get(id=pk)
            admin_profile = institution.user_object  # Get linked UserProfile
            
            # Delete the admin — this will auto-delete Institution due to CASCADE
            admin_profile.delete()

            return Response({"message": "Admin and Institution deleted"}, status=status.HTTP_204_NO_CONTENT)
        
        except Institution.DoesNotExist:
            return Response({"error": "Institution not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)  
    
class Institution_adminDelete(generics.DestroyAPIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[permissions.IsAdminUser]
    serializer_class=UserSerializer
    
    def get_queryset(self):
        return UserProfile.objects.filter(role="institution_admin")
    
    
class CreateListStaffView(generics.ListCreateAPIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperadminOrReadOnlyForStaff]  # Only admins can create staff
    serializer_class = UserSerializer

    def perform_create(self, serializer):
        user=serializer.save(role='staff')
        
        staff_role = self.request.data.get('staff_role', 'general_staff')  # default if not provided
        can_access_school_college = False
        can_access_academics = False
        can_access_package = False

        if staff_role == 'institution_manager':
            can_access_school_college = True
        elif staff_role == 'academic_manager':
            can_access_academics = True
        elif staff_role == 'subscription_manager':
            can_access_package = True

        # Create StaffRole object
        StaffRole.objects.create(
            user=user,
            staff_role=staff_role,
            can_access_school_college=can_access_school_college,
            can_access_academics=can_access_academics,
            can_access_package=can_access_package
        )


    def get_queryset(self):
        return UserProfile.objects.filter(role="staff")
    
class DeleteUpdateRetrieveStaffView(generics.RetrieveUpdateDestroyAPIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[permissions.IsAdminUser]
    serializer_class=UserSerializer
    
    def get_queryset(self):
        return UserProfile.objects.filter(role='staff')

class KeralaPincodeListView(APIView):
    def get(self, request):
        pincodes = [code for code, _ in PINCODE_CHOICES]
        return Response(pincodes, status=status.HTTP_200_OK)    

class CreateSchoolView(generics.ListCreateAPIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperAdminOrInstitutionManager]  # Only superadmin
    serializer_class = SchoolSerializer
    queryset = School.objects.all()
    
    def create(self, request, *args, **kwargs):
        print(request.data)  # 👈 This will show exactly what the frontend sends
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        institution_id = self.kwargs.get('pk')

        try:
            institution = Institution.objects.get(id=institution_id)

            if School.objects.filter(instution_obj=institution).exists():
                raise serializers.ValidationError("A school is already registered for this institution.")
            if College.objects.filter(instution_obj=institution).exists():
                raise serializers.ValidationError("This institution already has a college, cannot add school.")

            serializer.save(instution_obj=institution)

        except Institution.DoesNotExist:
            raise serializers.ValidationError("Institution not found.")

class UpdateRetireveDeleteSchoolView(generics.RetrieveUpdateDestroyAPIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[IsSuperAdminOrInstitutionManager]
    serializer_class = SchoolSerializer
    queryset = School.objects.all()
    
    
class GetSchoolByInstitution(APIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes = [IsSuperAdminOrInstitutionManager]

    def get(self, request, institution_id):
        try:
            school = School.objects.get(instution_obj__id=institution_id)
            serializer = SchoolSerializer(school)
            return Response(serializer.data)
        except School.DoesNotExist:
            return Response({"detail": "No School found for this institution"})
        
    def put(self,request,institution_id):
        
        try:
            school = School.objects.get(instution_obj__id=institution_id)
        except School.DoesNotExist:
            return Response({"detail": "School not found"})

        serializer = SchoolSerializer(school, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors)
        
class GetCollegeByInstitution(APIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes = [IsSuperAdminOrInstitutionManager]

    def get(self, request, institution_id):
        try:
            college = College.objects.get(instution_obj__id=institution_id)
            serializer = CollegeSerializer(college)
            return Response(serializer.data)
        except College.DoesNotExist:
            return Response({"detail": "No college found for this institution"})
        
    def put(self,request,institution_id):
        
        try:
            college = College.objects.get(instution_obj__id=institution_id)
        except College.DoesNotExist:
            return Response({"detail": "College not found"})

        serializer = CollegeSerializer(college, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors)
            
class CreateCollegeView(generics.ListCreateAPIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[IsSuperAdminOrInstitutionManager]
    serializer_class=CollegeSerializer
    queryset=College.objects.all()
    
    def perform_create(self, serializer):
        id = self.kwargs.get('pk')
        try:
            institution_obj = Institution.objects.get(id=id)
            
            
            
            # Check if this institution already has a school or college
            if College.objects.filter(instution_obj=institution_obj).exists():
                raise serializers.ValidationError("A college is already registered for this institution.")

            if School.objects.filter(instution_obj=institution_obj).exists():
                raise serializers.ValidationError("This institution already has a school, cannot add college.")

            serializer.save(instution_obj=institution_obj)
        except Institution.DoesNotExist:
            raise serializers.ValidationError("Institution not found.")

from superadmin_app.constants.std_codes import STD_CODE_CHOICES
class STDCodeListView(APIView):
    def get(self, request):
        # Convert tuple list to list of dicts for the frontend
        codes = [{"value": code, "label": code} for code, label in STD_CODE_CHOICES]

        return Response(codes)
    
    
class UpdateRetireveDeleteCollegeView(generics.RetrieveUpdateDestroyAPIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[IsSuperAdminOrInstitutionManager]
    serializer_class = CollegeSerializer
    queryset = College.objects.all()

# views.py



class AdminLoginView(APIView):
    def post(self, request):
        serializer = AdminLoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        username = serializer.validated_data['username']
        password = serializer.validated_data['password']

        user = authenticate(username=username, password=password)

        if user is None:
            return Response({'detail': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

        if  user.role !='staff' and not user.is_superuser:
            return Response({'detail': 'You are not authorized to login here'}, status=status.HTTP_403_FORBIDDEN)

        token, _ = Token.objects.get_or_create(user=user)

        return Response({
            'token': token.key,
            'user_id': user.id,
            'username': user.username,
            'is_superuser': user.is_superuser,
            'is_staff': user.is_staff
        }, status=status.HTTP_200_OK)


class InstitutionAdminLoginView(generics.GenericAPIView)  :
    serializer_class=InstitutionAdminLoginSerializer
    
    def post(self,request,*args,**kwargs):
        serializer=InstitutionAdminLoginSerializer(data=request.data)
        
        if serializer.is_valid():
            user=serializer.validated_data['user']
            institution_type = serializer.validated_data['institution_type']
            
            # Safely get or create token
            token, _ = Token.objects.get_or_create(user=user)

        
            return Response({"message":"Login Successful",
                             "token": token.key,  # ✅ This is what your frontend needs
                "institution_type": institution_type,  # Optional
                "user": UserSerializer(user).data
                })
        
        return Response(serializer.errors)
    def get(self, request, *args, **kwargs):
        """
        Return the authenticated user's data (only if logged in).
        """
        if request.user.is_authenticated:
            return Response({
                "user": UserSerializer(request.user).data
            })
        return Response({"detail": "Authentication credentials were not provided."}, status=401)
    
    
    
class CreateListPackage(generics.ListCreateAPIView):
    
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[IsSuperAdminOrPackageManager]
    serializer_class=Subscription_packageSerializer
    
    def get_queryset(self):
        institution_type = self.request.query_params.get('institution_type')  # e.g., ?institution_type=school
        queryset = SubscriptionPackage.objects.all()
        
        if institution_type:
            queryset = queryset.filter(institution_type=institution_type)
        return queryset

    
    
    def perform_create(self, serializer):
        return serializer.save(user_obj=self.request.user)

    
class UpdateRetrievePackage(generics.RetrieveUpdateAPIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[permissions.IsAdminUser]
    serializer_class=Subscription_packageSerializer
    queryset=SubscriptionPackage.objects.all()
    
    
class ListPackage(generics.ListAPIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[IsInstitutionAdmin]
    serializer_class=Subscription_packageSerializer
    
    def get_queryset(self):
        user = self.request.user

        # Get the institution linked to the institution admin
        institution = Institution.objects.get(user_object=user)

        # Check the institution type (school or college)
        is_school = School.objects.filter(instution_obj=institution).exists()
        is_college = College.objects.filter(instution_obj=institution).exists()

        if is_school:
            institution_type = 'school'
        elif is_college:
            institution_type = 'college'
        else:
            institution_type = None

        # Filter packages based on institution type
        if institution_type:
            return SubscriptionPackage.objects.filter(institution_type=institution_type)
        else:
            return SubscriptionPackage.objects.none()
    



    

class ListSchoolsCollegesView(APIView):
    def get(self, request):
        # Check if ?active=true is passed in the query params
        active_only = request.query_params.get('active') == 'true'

        # Base queryset
        school_queryset = School.objects.select_related('instution_obj__user_object')
        college_queryset = College.objects.select_related('instution_obj__user_object')

        # Apply filter if active_only is true
        if active_only:
            school_queryset = school_queryset.filter(is_active=True)
            college_queryset = college_queryset.filter(is_active=True)

        # Build school data
        school_data = []
        for school in school_queryset:
            user = school.instution_obj.user_object
            school_data.append({
                "name": school.school_name,
                "registration_id": school.registration_id,
                "type": "school",
                "is_active": school.is_active,
                "username": user.username,
                "admin_id":user.id,
                "email": user.email,
                "phone":school.phone_number,
                "id":school.instution_obj.id
            })

        # Build college data
        college_data = []
        for college in college_queryset:
            user = college.instution_obj.user_object
            college_data.append({
                "name": college.college_name,
                "registration_id": college.registration_id,
                "type": "college",
                "is_active": college.is_active,
                "username": user.username,
                "email": user.email,
                "phone":college.phone_number,
                "id":college.instution_obj.id
            })

        # Combine and return
        combined_data = school_data + college_data
        return Response(combined_data)
    
# class ListInstitutionAdminsWithDetails(APIView):
#     def get(self, request):
#         institutions = Institution.objects.select_related('user_object').all()
#         results = []

#         for institution in institutions:
#             admin = institution.user_object

#             school = School.objects.filter(instution_obj=institution).first()
#             college = College.objects.filter(instution_obj=institution).first()

#             institution_type = None
#             if school:
#                 institution_type = 'School'
#             elif college:
#                 institution_type = 'College'
#             else:
#                 institution_type = 'Not Registered'

#             data = {
#                 "admin_id": admin.id,
#                 "admin_username": admin.username,
#                 "admin_email": admin.email,
#                 "institution_id": institution.id,
#                 "institution_created": institution.created_date,
#                 "institution_updated": institution.updated_date,
#                 "institution_type": institution_type,
#                 "school": None,
#                 "college": None
#             }

#             if school:
#                 data["school"] = {
#                     "name": school.school_name,
#                     "registration_id": school.registration_id,
#                     "is_active": school.is_active,
#                     "phone": school.phone_number
#                 }

#             if college:
#                 data["college"] = {
#                     "name": college.college_name,
#                     "registration_id": college.registration_id,
#                     "is_active": college.is_active,
#                     "phone": college.phone_number,
                    
#                 }

#             results.append(data)

#         return Response(results)
        
        
class ListInstitutionAdminsWithDetails(APIView):
    def get(self, request):
        university_filter = request.GET.get('university', '').strip().lower()
        institution_name_filter = request.GET.get('institution_name', '').strip().lower()

        institutions = Institution.objects.select_related('user_object').all()

        # Filter by university via related College -> University
        if university_filter:
            matching_colleges = College.objects.filter(
                university__name__icontains=university_filter
            ).values_list('instution_obj_id', flat=True)
            institutions = institutions.filter(id__in=matching_colleges)

        # Filter by school or college name
        if institution_name_filter:
            # Get schools matching name filter
            matching_schools = School.objects.filter(
                school_name__icontains=institution_name_filter
            ).values_list('instution_obj_id', flat=True)

            # Get colleges matching name filter
            matching_colleges = College.objects.filter(
                college_name__icontains=institution_name_filter
            ).values_list('instution_obj_id', flat=True)

            # Combine institution ids from both
            institution_ids = set(matching_schools) | set(matching_colleges)

            # Filter institutions accordingly
            institutions = institutions.filter(id__in=institution_ids)

        results = []

        for institution in institutions:
            admin = institution.user_object

            school = School.objects.filter(instution_obj=institution).first()
            college = College.objects.filter(instution_obj=institution).first()

            institution_type = None
            if school:
                institution_type = 'School'
            elif college:
                institution_type = 'College'
            else:
                institution_type = 'Not Registered'

            data = {
                "admin_id": admin.id,
                "admin_username": admin.username,
                "admin_email": admin.email,
                "institution_id": institution.id,
                "institution_created": institution.created_date.isoformat() if institution.created_date else None,
                "institution_updated": institution.updated_date.isoformat() if institution.updated_date else None,
                "institution_type": institution_type,
                "school": None,
                "college": None
            }

            if school:
                data["school"] = {
                    "name": school.school_name,
                    "registration_id": school.registration_id,
                    "is_active": school.is_active,
                    "phone": school.phone_number,
                    "district": school.district,
                }

            if college:
                data["college"] = {
                    "name": college.college_name,
                    "registration_id": college.registration_id,
                    "is_active": college.is_active,
                    "phone": college.phone_number,
                    "district": college.district,
                }
                if college.university:
                    data["college"]["university"] = {
                        "id": college.university.id,
                        "name": college.university.name,
                    }

            results.append(data)

        return Response(results)

import uuid

class CheckoutView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, *args, **kwargs):
        package_id = kwargs.get("pk")
        coupon_code = request.data.get("coupon_code")

        if not package_id:
            return Response({"error": "Package ID is required"}, status=400)

        # Fetch package
        try:
            package_obj = SubscriptionPackage.objects.get(id=package_id)
        except SubscriptionPackage.DoesNotExist:
            return Response({"error": "Package not found"}, status=404)

        # -------------------------------
        # Calculate upgrade or full price
        # -------------------------------
        old_payment = Payment.objects.filter(
            user_obj=request.user, is_paid=True
        ).order_by('-end_date').first()

        if old_payment and old_payment.end_date > timezone.now().date():
            # Upgrade calculation
            total_days = (old_payment.end_date - old_payment.start_date).days
            days_used = (timezone.now().date() - old_payment.start_date).days
            days_remaining = total_days - days_used
            old_price_per_day = old_payment.package_obj.price / total_days
            remaining_value = old_price_per_day * days_remaining
            price_to_pay = package_obj.price - remaining_value
        else:
            # Normal full price
            price_to_pay = package_obj.price

        price_to_pay = round(max(price_to_pay, 0), 2)

        # -------------------------------
        # Apply coupon (if provided)
        # -------------------------------
        coupon = None
        discount_applied = 0
        if coupon_code:
            try:
                coupon = Coupon.objects.get(code=coupon_code, is_active=True)
                old_pkg = old_payment.package_obj if old_payment else None
                print("Coupon code received:", coupon_code)
                if coupon.is_valid_for_upgrade(old_package=old_pkg, new_package=package_obj):
                    print("Coupon is valid, applying discount")
                    if coupon.discount_type == "fixed":
                        discount_applied = Decimal(coupon.discount_value)
                    elif coupon.discount_type == "percentage":
                        discount_applied = (price_to_pay * Decimal(coupon.discount_value)) / Decimal(100)

                    # Apply discount
                    price_to_pay = (price_to_pay - discount_applied).quantize(Decimal("0.01"))
                else:
                    return Response({"error": "Coupon is not valid for this upgrade"}, status=400)
            except Coupon.DoesNotExist:
                return Response({"error": "Invalid coupon code"}, status=400)
            
        print("Final price_to_pay (after coupon):", price_to_pay)

        # -------------------------------
        # Create Razorpay Order
        # -------------------------------
        client = razorpay.Client(auth=(KEY_ID, SECRET_KEY))
        data = {
            "amount": int(price_to_pay * 100),  # Razorpay expects amount in paise
            "currency": "INR",
            "receipt": f"order_rcptid_{request.user.id}_{package_id}"
        }
        payment = client.order.create(data=data)
        
        print("Razorpay order created:", payment)

        # -------------------------------
        # Save Pending Payment
        # -------------------------------
        PendingPayment.objects.create(
            user=request.user,
            package_obj=package_obj,
            coupon=coupon,            # store Coupon object
            amount=price_to_pay,      # final amount after discount
            discount_applied=discount_applied,  # store applied discount
            razorpay_order_id=payment["id"],
            is_paid=False
        )

        # -------------------------------
        # Return response to frontend
        # -------------------------------
        payment['package_id'] = package_id
        payment['amount_to_pay'] = price_to_pay
        payment['discount_applied'] = discount_applied
        payment['coupon_code'] = coupon_code

        return Response(data=payment, status=200)



    # def post(self, request, *args, **kwargs):
    #     package_id = kwargs.get("pk")
    #     if not package_id:
    #         return Response({"error": "Package ID is required"}, status=400)

    #     package_obj = SubscriptionPackage.objects.get(id=package_id)
    #     price = int(float(package_obj.price) * 100)  # Razorpay expects amount in paise

    #     # Check if user has an existing payment
    #     old_payment = Payment.objects.filter(user_obj=request.user, is_paid=True).order_by('-end_date').first()
    #     if old_payment and old_payment.end_date > timezone.now().date():
    #         # Calculate upgrade amount
    #         total_days = (old_payment.end_date - old_payment.start_date).days
    #         days_used = (timezone.now().date() - old_payment.start_date).days
    #         days_remaining = total_days - days_used
    #         old_price_per_day = old_payment.package_obj.price / total_days
    #         remaining_value = old_price_per_day * days_remaining
    #         price_to_pay = package_obj.price - remaining_value
    #         price_to_pay = round(max(price_to_pay, 0), 2)
    #     else:
    #         price_to_pay = package_obj.price

    #     # Razorpay payment
    #     client = razorpay.Client(auth=(KEY_ID, SECRET_KEY))
    #     data = {"amount": int(price_to_pay * 100), "currency": "INR", "receipt": ""}
    #     payment = client.order.create(data=data)
    #     payment['package_id'] = package_id
    #     payment['amount_to_pay'] = price_to_pay  # send pro-rated amount to frontend

    #     return Response(data=payment)
        
       


from django.utils import timezone

# class PaymentVerifyView(APIView):
#     authentication_classes = [authentication.TokenAuthentication]
#     permission_classes = [permissions.IsAuthenticated]

#     def post(self, request, *args, **kwargs):
#         razorpay_payment_id = request.data.get("razorpay_payment_id")
#         razorpay_order_id = request.data.get("razorpay_order_id")
#         razorpay_signature = request.data.get("razorpay_signature")
#         package_id = request.data.get("package_id")
#         coupon_code = request.data.get("coupon_code")  # optional
#         amount_to_pay = request.data.get("amount_to_pay")  # 👈 carry this from checkout

#         if not (razorpay_payment_id and razorpay_order_id and razorpay_signature):
#             return Response({"error": "Missing payment details"}, status=400)

#         try:
#             package_obj = SubscriptionPackage.objects.get(id=package_id)
#         except SubscriptionPackage.DoesNotExist:
#             return Response({"error": "Package not found"}, status=404)

#         # ---------- STEP 1: Verify signature with Razorpay ----------
#         client = razorpay.Client(auth=(KEY_ID, SECRET_KEY))
#         params_dict = {
#             "razorpay_order_id": razorpay_order_id,
#             "razorpay_payment_id": razorpay_payment_id,
#             "razorpay_signature": razorpay_signature,
#         }

#         try:
#             client.utility.verify_payment_signature(params_dict)
#         except razorpay.errors.SignatureVerificationError:
#             return Response({"error": "Payment verification failed"}, status=400)

#         # ---------- STEP 2: Final amount (handle coupon again for safety) ----------
#         final_amount = float(amount_to_pay) if amount_to_pay else float(package_obj.price)

#         if coupon_code:
#             try:
#                 coupon = Coupon.objects.get(code=coupon_code, is_active=True)
#                 discount = (final_amount * coupon.discount_percent) / 100
#                 final_amount -= discount
#                 coupon.is_active = False
#                 coupon.save()
#             except Coupon.DoesNotExist:
#                 pass

#         # ---------- STEP 3: Mark payment as success ----------
#         payment = Payment.objects.create(
#             user_obj=request.user,
#             package_obj=package_obj,
#             razorpay_payment_id=razorpay_payment_id,
#             razorpay_order_id=razorpay_order_id,
#             razorpay_signature=razorpay_signature,
#             amount=round(final_amount, 2),   # 👈 save the actual paid amount
#             is_paid=True,
#             coupon_code=coupon_code if coupon_code else None,
#         )

#         # ---------- STEP 4: Send notification ----------
#         Notification.objects.create(
#             user=request.user,
#             title="Payment Successful",
#             message=f"Your payment of ₹{final_amount} for {package_obj.name} has been verified successfully.",
#             type="payment"
#         )

#         return Response({"success": True, "message": "Payment verified successfully"})


class PaymentVerifyView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, *args, **kwargs):
        client = razorpay.Client(auth=(KEY_ID, SECRET_KEY))
        razorpay_order_id = request.data.get("razorpay_order_id")
        package_id = request.data.get("package_id")

        if not package_id:
            return Response({"error": "Package ID is required"}, status=400)

        try:
            package_obj = SubscriptionPackage.objects.get(id=package_id)
        except SubscriptionPackage.DoesNotExist:
            return Response({"error": "Package not found"}, status=404)

        # Verify Razorpay signature
        params_dict = {
            "razorpay_order_id": request.data.get("razorpay_order_id"),
            "razorpay_payment_id": request.data.get("razorpay_payment_id"),
            "razorpay_signature": request.data.get("razorpay_signature"),
        }

        try:
            client.utility.verify_payment_signature(params_dict)
        except razorpay.errors.SignatureVerificationError:
            return Response({"error": "Payment verification failed"}, status=400)

        # Get PendingPayment
        pending_payment = PendingPayment.objects.filter(
            razorpay_order_id=razorpay_order_id,
            user=request.user,
            package_obj=package_obj,
            is_paid=False
        ).first()

        if not pending_payment:
            return Response({"error": "No pending payment found"}, status=400)

        now_date = timezone.now().date()
        coupon = pending_payment.coupon

        # Apply discount
        if coupon:
            if coupon.discount_type == "percentage":
                discount_amount = (package_obj.price * coupon.discount_value) / 100
            else:
                discount_amount = coupon.discount_value
            price_to_pay = package_obj.price - discount_amount
        else:
            price_to_pay = package_obj.price

        institution = Institution.objects.get(user_object=request.user)
        # Determine actual institution name
        school = School.objects.filter(instution_obj=institution).first()
        college = College.objects.filter(instution_obj=institution).first()

        if school:
            inst_name = school.school_name  # or school.school_name depending on your model
        elif college:
            inst_name = college.college_name  # or college.college_name
        else:
            inst_name = "Institution"
        
        # -------------------------------
        # Handle upgrade or normal
        # -------------------------------
        latest_payment = Payment.objects.filter(
            user_obj=request.user, is_paid=True
        ).order_by('-end_date').first()

        institution = Institution.objects.get(user_object=request.user)

        if latest_payment and latest_payment.end_date > now_date:
            # Upgrade logic ...
            total_days = (latest_payment.end_date - latest_payment.start_date).days
            days_used = (now_date - latest_payment.start_date).days
            days_remaining = total_days - days_used
            old_price_per_day = latest_payment.package_obj.price / total_days
            remaining_value = old_price_per_day * days_remaining
            upgrade_amount = price_to_pay - remaining_value
            upgrade_amount = round(max(upgrade_amount, 0), 2)

            final_amount = upgrade_amount
            # ✅ Convert Decimal → float
            if isinstance(upgrade_amount, Decimal):
                upgrade_amount = float(upgrade_amount)

            # ✅ Save upgrade history safely
            upgrade_entry = {
                "old_package": str(latest_payment.package_obj.package),
                "new_package": str(package_obj.package),
                "paid_amount": float(upgrade_amount),     # ensure float
                "date": str(now_date),                    # ensure string
                "coupon_code": coupon.code if coupon else None
            }

            history = latest_payment.upgrade_history or []
            history.append(upgrade_entry)
            latest_payment.upgrade_history = history

            # Update payment record
            latest_payment.package_obj = package_obj
            latest_payment.end_date = latest_payment.start_date + timedelta(days=365)
            latest_payment.final_amount = float(upgrade_amount)  # also cast here
            latest_payment.coupon_code = coupon.code if coupon else None
            latest_payment.is_paid = True
            latest_payment.save()

            # History
            PaymentHistory.objects.create(
                institution=institution,
                payment=latest_payment,
                original_price=package_obj.price,
                final_amount=upgrade_amount,  # ✅ amount actually paid for upgrade
                coupon_code=coupon.code if coupon else None,
                package=package_obj.package
            )

            payment_type = "upgrade"

        else:
            
            final_amount = price_to_pay  
            # Normal new payment
            activation_date = timezone.now()
            latest_payment = Payment.objects.create(
                user_obj=request.user,
                package_obj=package_obj,
                start_date=activation_date,
                end_date=activation_date + timedelta(days=365),
                is_paid=True,
                final_amount=price_to_pay,
                coupon_code=coupon.code if coupon else None
            )

            PaymentHistory.objects.create(
                institution=institution,
                payment=latest_payment,
                original_price=package_obj.price,
                final_amount=price_to_pay,
                coupon_code=coupon.code if coupon else None,
                package=package_obj.package
            )

            payment_type = "normal"

        # ✅ Mark PendingPayment as paid
        pending_payment.is_paid = True
        pending_payment.save()

        #STEP 4: Send notification ----------
       # after payment is verified & saved
        Notification.objects.create(
            institution=institution,
            notification_type="payment_done",
            title="Payment Successful",
            message=f"Payment of ₹{latest_payment.final_amount} for {inst_name} was completed successfully.",
            is_read=False
        )

        # clean up old "payment_due" or "trial_expired" since they are no longer relevant
        Notification.objects.filter(
            institution=institution,
            notification_type__in=["trial_expired", "payment_due"]
        ).delete()

        # ✅ Final Response
        return Response({
            "message": "Payment verified successfully",
            "payment_id": latest_payment.id,
            "amount_paid": float(latest_payment.final_amount),
            "type": payment_type
        }, status=200)




class LatestPaymentReportView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperadminOrReadOnlyForStaff]

    def get(self, request, institution_id):
        try:
            # Get institution object
            institution = get_object_or_404(Institution, id=institution_id)

            # Get the admin user associated with the institution
            user = UserProfile.objects.filter(
                institution=institution, role="institution_admin"
            ).first()
            if not user:
                return Response(
                    {"detail": "Institution admin user not found."}, status=404
                )

            # Get the latest paid payment
            latest_payment = (
                Payment.objects.filter(user_obj=user, is_paid=True)
                .select_related("user_obj", "package_obj")
                .order_by("-start_date")
                .first()
            )

            # Get school/college info
            school = School.objects.filter(instution_obj=institution).first()
            college = College.objects.filter(instution_obj=institution).first()

            # If no paid payment exists → handle trial
            if not latest_payment:
                institution_type, institution_name, registration_id, created_date = None, None, None, None

                if school:
                    institution_type = "school"
                    institution_name = school.school_name
                    registration_id = school.registration_id
                    created_date = school.created_date
                elif college:
                    institution_type = "college"
                    institution_name = college.college_name
                    registration_id = college.registration_id
                    created_date = college.created_date
                else:
                    return Response({"detail": "Institution not found."}, status=404)

                trial_expiry = created_date + timedelta(days=7)
                if timezone.now().date() <= trial_expiry:
                    return Response(
                        {
                            "institution_id": institution.id,
                            "institution_type": institution_type,
                            "institution_name": institution_name,
                            "registration_id": registration_id,
                            "username": user.username,
                            "email": user.email,
                            "payment_status": "trial",
                            "start_date": created_date,
                            "end_date": trial_expiry,
                            "is_paid": False,
                            "upgrade_history": [],
                            "total_amount_paid": 0.0,
                        },
                        status=200,
                    )

                return Response(
                    {"detail": "No paid payment found and trial expired."}, status=404
                )

            # ✅ Calculate total paid using aggregation
            total_paid = (
                Payment.objects.filter(user_obj=user, is_paid=True)
                .aggregate(total=Sum("final_amount"))
                .get("total") or 0
            )

            # ----- Prepare Response Data -----
            data = {
                "institution_id": institution.id,
                "institution_type": "school" if school else "college" if college else "unknown",
                "institution_name": (
                    school.school_name if school else college.college_name if college else "N/A"
                ),
                "registration_id": (
                    school.registration_id if school else college.registration_id if college else "N/A"
                ),
                "username": user.username,
                "email": user.email,
                "package": latest_payment.package_obj.package,
                "coupon_code": latest_payment.coupon_code if latest_payment.coupon_code else None,
                "final_amount": latest_payment.final_amount,
                "original_price": latest_payment.package_obj.price,
                "start_date": latest_payment.start_date,
                "end_date": latest_payment.end_date,
                "is_paid": latest_payment.is_paid,
                "upgrade_history": latest_payment.upgrade_history,
                "total_amount_paid": float(total_paid),  # ✅ Now shows ALL paid total
            }

            return Response(data, status=200)

        except Exception as e:
            return Response({"detail": "Error occurred", "error": str(e)}, status=400)

# class LatestPaymentReportView(APIView):
#     authentication_classes = [authentication.TokenAuthentication]
#     permission_classes = [permissions.IsAdminUser]

#     def get(self, request, institution_id):
#         try:
#             # Get institution object
#             institution = get_object_or_404(Institution, id=institution_id)

#             # Get the admin user associated with the institution
#             user = UserProfile.objects.filter(
#                 institution=institution, role="institution_admin"
#             ).first()
#             if not user:
#                 return Response(
#                     {"detail": "Institution admin user not found."}, status=404
#                 )

#             # Get the latest paid payment
#             latest_payment = (
#                 Payment.objects.filter(user_obj=user, is_paid=True)
#                 .select_related("user_obj", "package_obj")
#                 .order_by("-start_date")
#                 .first()
#             )
#             total_amount_paid = latest_payment.final_amount or 0
#             # Get school/college info
#             school = School.objects.filter(instution_obj=institution).first()
#             college = College.objects.filter(instution_obj=institution).first()

#             if not latest_payment:
#                 # Default fallback values
#                 institution_type = None
#                 institution_name = None
#                 registration_id = None
#                 created_date = None

#                 if school:
#                     institution_type = "school"
#                     institution_name = school.school_name
#                     registration_id = school.registration_id
#                     created_date = school.created_date
#                 elif college:
#                     institution_type = "college"
#                     institution_name = college.college_name
#                     registration_id = college.registration_id
#                     created_date = college.created_date
#                 else:
#                     return Response({"detail": "Institution not found."}, status=404)

#                 # Add all upgrade amounts from upgrade_history
#                 for upgrade in latest_payment.upgrade_history:
#                     total_amount_paid += upgrade.get("upgrade_amount", 0)
                
#                 # Trial logic
#                 trial_expiry = created_date + timedelta(days=7)  # your trial duration
#                 if timezone.now() <= trial_expiry:
#                     return Response(
#                         {
#                             "institution_id": institution.id,
#                             "institution_type": institution_type,
#                             "institution_name": institution_name,
#                             "registration_id": registration_id,
#                             "username": user.username,
#                             "email": user.email,
#                             "payment_status": "trial",
#                             "start_date": created_date,
#                             "end_date": trial_expiry,
#                             "is_paid": False,
#                             "upgrade_history": [],  # No upgrade history in trial
#                         },
#                         status=200,
#                     )

#                 return Response(
#                     {"detail": "No paid payment found and trial expired."}, status=404
#                 )

#                 # Paid user details
#             data = {
#                 "institution_id": institution.id,
#                 "institution_type": "school" if school else "college" if college else "unknown",
#                 "institution_name": school.school_name if school else college.college_name if college else "N/A",
#                 "registration_id": school.registration_id if school else college.registration_id if college else "N/A",
#                 "username": user.username,
#                 "email": user.email,
#                 "package": latest_payment.package_obj.package,

#                 # ✅ Show coupon info & pricing details
#                 "coupon_code": latest_payment.coupon_code if latest_payment.coupon_code else None,
#                 "final_amount": (
#                     latest_payment.final_amount 
#                     if latest_payment.final_amount is not None 
#                     else latest_payment.package_obj.price
#                 ),
#                 "original_price": latest_payment.package_obj.price,

#                 "start_date": latest_payment.start_date,
#                 "end_date": latest_payment.end_date,
#                 "is_paid": latest_payment.is_paid,
#                 "upgrade_history": latest_payment.upgrade_history,
#                 "total_amount_paid" : total_amount_paid
#             }




#             return Response(data, status=200)

#         except Exception as e:
#             return Response({"detail": "Error occurred", "error": str(e)}, status=400)


from django.db.models import Sum

class InstitutionTotalPaymentView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)

        total_paid = PaymentHistory.objects.filter(institution=institution).aggregate(
            total=Sum("final_amount")
        )["total"] or 0

        return Response({
            "institution_id": institution.id,
            "institution_name": getattr(institution, "name", f"Institution {institution.id}"),
            "total_paid": float(total_paid)
        })


from django.db.models.functions import TruncMonth,TruncDay   #It truncates a datetime/date field to just the day (removes hours, minutes, seconds).

class PaymentHistoryView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperadminOrStaff]

    def get(self, request):
        type = request.GET.get("type", "daily")

        # use paid_at as the transaction date
        date_field = "paid_at"

        if type == "monthly":
            payments = (
                PaymentHistory.objects.annotate(month=TruncMonth(date_field))
                .values("month")
                .annotate(total=Sum("final_amount"))  # use actual paid amount
                .order_by("month")
            )
            data = [
                {"date": p["month"].strftime("%Y-%m"), "total": float(p["total"])}
                for p in payments
            ]

        else:  # daily
            payments = (
                PaymentHistory.objects.annotate(day=TruncDay(date_field))
                .values("day")
                .annotate(total=Sum("final_amount"))
                .order_by("day")
            )
            data = [
                {"date": p["day"].strftime("%Y-%m-%d"), "total": float(p["total"])}
                for p in payments
            ]

        return Response(data, status=200)

    
from datetime import date

class SubscriptionUpgradeCalculator:
    """
    Handles pro-rated subscription upgrade cost calculation.
    """

    def __init__(self, old_payment, new_package_price):
        self.old_payment = old_payment
        self.new_package_price = new_package_price

    def calculate(self):
        total_days = (self.old_payment.end_date - self.old_payment.start_date).days
        days_used = (date.today() - self.old_payment.start_date).days
        days_remaining = total_days - days_used

        if days_remaining <= 0:
            return self.new_package_price  # Expired — pay full price

        # Calculate remaining value of old subscription
        old_price_per_day = self.old_payment.package_obj.price / total_days
        remaining_old_value = old_price_per_day * days_remaining

        # Upgrade amount = new package price - remaining old value
        upgrade_amount = self.new_package_price - remaining_old_value
        return round(max(upgrade_amount, 0), 2)  # Round to 2 decimals

class UpgradePreviewView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            new_package_id = self.kwargs.get('package_id')
            new_package = SubscriptionPackage.objects.get(id=new_package_id)

            # Get latest active payment
            latest_payment = Payment.objects.filter(
                user_obj=request.user,
                is_paid=True
            ).order_by('-end_date').first()

            if not latest_payment or latest_payment.end_date <= timezone.now().date():
                # No active subscription or expired → full payment
                return Response({
                    "is_upgrade": False,
                    "amount_to_pay": float(new_package.price),
                    "discount_amount": 0,
                    "final_amount_to_pay": float(new_package.price),
                    "message": "No active subscription found or subscription expired. Full price applies."
                })

            # Check if it's actually an upgrade
            if new_package.price <= latest_payment.package_obj.price:
                return Response({
                    "is_upgrade": False,
                    "amount_to_pay": 0,
                    "discount_amount": 0,
                    "final_amount_to_pay": 0,
                    "message": "Selected package is not higher than the current one."
                })

            # Calculate pro-rated upgrade cost
            calc = SubscriptionUpgradeCalculator(latest_payment, Decimal(new_package.price))
            upgrade_amount = calc.calculate()

            # Check coupon (from query params)
            coupon_code = request.query_params.get("coupon_code")
            discount_amount = 0
            if coupon_code:
                try:
                    coupon = Coupon.objects.get(code=coupon_code, is_active=True)
                    # Validate coupon matches upgrade path
                    if (coupon.from_package == latest_payment.package_obj and
                        coupon.to_package.id == new_package.id):
                        if coupon.discount_type == "fixed":
                            discount_amount = coupon.discount_value
                        elif coupon.discount_type == "percentage":
                            discount_amount = (upgrade_amount * coupon.discount_value) / 100
                except Coupon.DoesNotExist:
                    discount_amount = 0  # Invalid coupon ignored

            final_amount_to_pay = max(upgrade_amount - discount_amount, 0)

            return Response({
                "is_upgrade": True,
                "old_package": latest_payment.package_obj.package,
                "new_package": new_package.package,
                "amount_to_pay": float(upgrade_amount),
                "discount_amount": float(discount_amount),
                "final_amount_to_pay": float(final_amount_to_pay),
                "coupon_code": coupon_code if coupon_code else None,
                "message": f"Upgrade from {latest_payment.package_obj.package} to {new_package.package}"
            })

        except SubscriptionPackage.DoesNotExist:
            return Response({"error": "Package not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=400)

class DeactivateView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        today = timezone.now()

        try:
            institution = user.institution
        except Institution.DoesNotExist:
            return Response({"error": "Institution not found"})

        latest_payment = Payment.objects.filter(user_obj=user, is_paid=True).order_by('-end_date').first()

        # School check
        school = School.objects.filter(instution_obj=institution).first()
        if school:
            trial_end = school.created_date + timedelta(minutes=4)  # change to days=30 for real
            if today > trial_end and not latest_payment:
                school.is_active = False
                school.save()

                Notification.objects.get_or_create(
                    institution=institution,
                    notification_type='trial_expired',
                    title="Trial Expired",
                    message=f"The trial period for school '{school.school_name}' has expired.",
                    is_read=False
                )

                return Response({"trial_status": "expired", "type": "school"}, status=200)

            if school.activation_date and school.is_active and latest_payment:
                Notification.objects.filter(
                    institution=institution,
                    notification_type__in=['trial_expired', 'payment_due']
                ).delete()

                expiry_date = school.activation_date + timedelta(days=365)  # yearly only
                if today > expiry_date:
                    school.is_active = False
                    school.save()

                    Notification.objects.get_or_create(
                        institution=institution,
                        notification_type='payment_due',
                        title="Subscription Expired",
                        message=f"The subscription for school '{school.school_name}' has expired.",
                        is_read=False
                    )

                    return Response({"trial_status": "expired", "type": "school", "reason": "subscription_expired"}, status=200)

        # College check
        college = College.objects.filter(instution_obj=institution).first()
        if college:
            trial_end = college.created_date + timedelta(minutes=4)  # change to days=30 for real
            if today > trial_end and not latest_payment:
                college.is_active = False
                college.save()

                Notification.objects.get_or_create(
                    institution=institution,
                    notification_type='trial_expired',
                    title="Trial Expired",
                    message=f"The trial period for college '{college.college_name}' has expired.",
                    is_read=False
                )

                return Response({"trial_status": "expired", "type": "college"}, status=200)

            if college.activation_date and college.is_active and latest_payment:
                Notification.objects.filter(
                    institution=institution,
                    notification_type__in=['trial_expired', 'payment_due']
                ).delete()

                expiry_date = college.activation_date + timedelta(days=365)  # yearly only
                if today > expiry_date:
                    college.is_active = False
                    college.save()

                    Notification.objects.get_or_create(
                        institution=institution,
                        notification_type='payment_due',
                        title="Subscription Expired",
                        message=f"The subscription for college '{college.college_name}' has expired.",
                        is_read=False
                    )

                    return Response({"trial_status": "expired", "type": "college", "reason": "subscription_expired"}, status=200)

        # Check if already deactivated
        if school and not school.is_active:
            return Response({"trial_status": "expired", "type": "school", "reason": "already_deactivated"}, status=200)
        if college and not college.is_active:
            return Response({"trial_status": "expired", "type": "college", "reason": "already_deactivated"}, status=200)

        unread_count = Notification.objects.filter(is_read=False).count()
        print("🔴 Unread Notification Count:", unread_count)

        return Response({"trial_status": "valid", "unread_notifications_global": unread_count}, status=200)


@api_view(['POST'])

def mark_all_notifications_read(request):
    Notification.objects.filter(is_read=False).update(is_read=True)
    return Response({'status': 'All notifications marked as read'})

class NotificationListView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperadminOrStaff]

    def get(self, request):
        

        notifications = Notification.objects.all().order_by('-created_at')[:50]  # latest 50
        serializer = NotificationSerializer(notifications, many=True)
        return Response(serializer.data)

    
    
class Institution_HomepageView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        institution = user.institution

        school = institution.school_set.first()
        college = institution.college_set.first()

        if school:
            institution_type = "school"
            institution_name = school.school_name
            homepage_message = "Welcome to the School Admin Homepage!"
        elif college:
            institution_type = "college"
            institution_name = college.college_name
            homepage_message = "Welcome to the College Admin Homepage!"
        else:
            institution_type = "unknown"
            institution_name = "Unknown"
            homepage_message = "Institution type not recognized."

        return Response({
            "message": f"Welcome, {user.username}!",
            "role": user.role,
            "institution_type": institution_type,
            "institution_name": institution_name,
            "homepage": homepage_message
        })


from django.contrib.auth import get_user_model
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import send_mail
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

User = get_user_model()

class PasswordResetRequestView(APIView):
    def post(self, request):
        email = request.data.get('email')
        if not email:
            return Response({"detail": "Email is required."}, status=400)

        try:
            user = User.objects.get(email=email, role='institution_admin')
        except User.DoesNotExist:
            return Response({"detail": "No institution admin with this email."}, status=404)

        token = default_token_generator.make_token(user)
        reset_link = f"http://localhost:5173/reset-password/{token}?email={email}"

        send_mail(
            subject='Reset Your Password',
            message=f'Click the link to reset your password: {reset_link}',
            from_email='noreply@yourapp.com',
            recipient_list=[email],
            fail_silently=False,
        )

        return Response({"detail": "Reset link sent."}, status=200)



from django.contrib.auth import get_user_model
from django.contrib.auth.tokens import default_token_generator
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

User = get_user_model()

class PasswordResetConfirmView(APIView):
    def post(self, request):
        print("DATA:", request.data)
        email = request.data.get('email')
        token = request.data.get('token')
        password = request.data.get('password')

        if not email or not token or not password:
            return Response({"detail": "Missing fields."}, status=400)

        try:
            user = User.objects.get(email=email, role='institution_admin')
        except User.DoesNotExist:
            return Response({"detail": "Invalid email or role."}, status=404)

        if not default_token_generator.check_token(user, token):
            return Response({"detail": "Invalid or expired token."}, status=400)

        user.set_password(password)
        user.save()

        return Response({"detail": "Password reset successful."}, status=200)




class TotalInstitutionCountView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperadminOrStaff]

    def get(self, request):
        school_count = School.objects.count()
        college_count = College.objects.count()
        total_institution_count = school_count + college_count
        
        active_school_count = School.objects.filter(is_active=True).count()
        inactive_school_count= School.objects.filter(is_active=False).count()
        active_college_count = College.objects.filter(is_active=True).count()
        inactive_college_count=College.objects.filter(is_active=False).count()
        total_inactive_institution_count=inactive_school_count+inactive_college_count
        total_active_institution_count = active_school_count + active_college_count
        
        total_amount_paid = PaymentHistory.objects.aggregate(
            total_amount=Sum('final_amount')
        )['total_amount'] or 0

        
        staff_count=User.objects.filter(role='staff').count()

        return Response({
            "total_schools": school_count,
            "total_colleges": college_count,
            "total_institutions": total_institution_count,
            "active_school_count":active_school_count,
            "active_college_count":active_college_count,
            "total_active_institution_count":total_active_institution_count,
            'inactive_school_count':inactive_school_count,
            'inactive_college_count':inactive_college_count,
            'total_inactive_institution_count':total_inactive_institution_count,
            'total_amount':float(total_amount_paid),
            'staff_count':staff_count
        })
        
    
class InstitutionDetailView(APIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperadminOrStaff]

    def get(self, request, institution_id):
        try:
            institution = Institution.objects.get(id=institution_id)
            user = institution.user_object  # This gives username/email

            school = School.objects.filter(instution_obj=institution).first()
            college = College.objects.filter(instution_obj=institution).first()

            if school:
                school_data = SchoolSerializer(school).data
                school_data['type'] = 'school'
                school_data['username'] = user.username
                school_data['email'] = user.email
                return Response(school_data)

            elif college:
                college_data = CollegeSerializer(college).data
                college_data['type'] = 'college'
                college_data['username'] = user.username
                college_data['email'] = user.email
                return Response(college_data)

            else:
                return Response({"detail": "No school or college found for this institution"})

        except Institution.DoesNotExist:
            return Response({"detail": "Institution not found"})


class DeactivateInstitutionView(APIView):
    def post(self, request, institution_type, institution_id):
        if institution_type == "school":
            try:
                school = School.objects.get(id=institution_id)
                school.is_active = False
                school.save()
                return Response({"message": "School deactivated successfully"}, status=status.HTTP_200_OK)
            except School.DoesNotExist:
                return Response({"error": "School not found"}, status=status.HTTP_404_NOT_FOUND)

        elif institution_type == "college":
            try:
                college = College.objects.get(id=institution_id)
                college.is_active = False
                college.save()
                return Response({"message": "College deactivated successfully"}, status=status.HTTP_200_OK)
            except College.DoesNotExist:
                return Response({"error": "College not found"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"error": "Invalid institution type"}, status=status.HTTP_400_BAD_REQUEST)

class UniversityView(generics.ListCreateAPIView)  :
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[IsSuperAdminOrAcademicManager]
    queryset = University.objects.all()
    serializer_class = UniversitySerializer
    
class ListUniversityView(generics.ListAPIView)  :
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[IsSuperAdminOrAcademicManager]
    queryset = University.objects.all()
    serializer_class = UniversityNestedSerializer
    
class UniversityDetailView(generics.RetrieveAPIView):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsSuperAdminOrAcademicManager]
    queryset = University.objects.all()
    serializer_class = UniversityNestedSerializer
        


class CourseView(generics.ListCreateAPIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[IsSuperAdminOrAcademicManager]

    serializer_class = CourseSerializer

    def get_queryset(self):
            university_id = self.request.query_params.get('university_id')
            if university_id:
                return Course.objects.filter(university_id=university_id)
            return Course.objects.all()


import openpyxl
from rest_framework.parsers import MultiPartParser    # handle file uploads
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

@method_decorator(csrf_exempt, name='dispatch')    
# class BulkCourseUploadView(APIView):
#     parser_classes = [MultiPartParser]
#     permission_classes = [permissions.IsAuthenticated]

#     def post(self, request):
#         file = request.FILES.get('file')
#         if not file:
#             return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             workbook = openpyxl.load_workbook(file)
#             sheet = workbook.active

#             for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
#                 name = row[0]
#                 if name:
#                     Course.objects.get_or_create(name=name)

#             return Response({"message": "Courses uploaded successfully!"})
#         except Exception as e:
#             return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class BulkCourseUploadView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsSuperAdminOrAcademicManager]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                course_name = row[0]
                university_name = row[1]

                if course_name and university_name:
                    # Get or create the university
                    university, _ = University.objects.get_or_create(name=university_name)

                    # Create the course under the university
                    Course.objects.get_or_create(name=course_name, university=university)

            return Response({"message": "Courses uploaded successfully!"})
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DepartmentView(generics.ListCreateAPIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[IsSuperAdminOrAcademicManager]
   
    serializer_class = DepartmentSerializer
    
    def get_queryset(self):
        course_id = self.request.query_params.get('course')
        university_id = self.request.query_params.get('university')

        if course_id:
            return Department.objects.filter(course_id=course_id)
        elif university_id:
            return Department.objects.filter(course__university_id=university_id)

        return Department.objects.all()

class SemesterView(generics.ListCreateAPIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[IsSuperAdminOrAcademicManager]
    queryset = Semester.objects.all()
    serializer_class = SemesterSerializer
    
    def get_queryset(self):
        university_id = self.request.query_params.get('university')
        course_id = self.request.query_params.get('course')
        dept_id = self.request.query_params.get('department')

        if course_id and dept_id:
            return Semester.objects.filter(course_id=course_id, department_id=dept_id)
        elif course_id:
            return Semester.objects.filter(course_id=course_id)
        elif university_id:
            return Department.objects.filter(course__university_id=university_id)
        return Semester.objects.all()

class SubjectView(generics.ListCreateAPIView):
    authentication_classes=[authentication.TokenAuthentication]
    permission_classes=[IsSuperAdminOrAcademicManager]
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    
    
class BulkSubjectUploadView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsSuperAdminOrAcademicManager]

    def post(self, request):
        file = request.FILES.get('file')
        semester_id = request.data.get('semester')

        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        if not semester_id:
            return Response({"error": "Semester ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            semester = Semester.objects.get(id=semester_id)

            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header
                code = row[0]
                name = row[1]
                if code and name:
                    Subject.objects.get_or_create(
                        name=name.strip(),
                        semester=semester,
                        defaults={'code': code.strip() if code else None}
                    )

            return Response({"message": "Subjects uploaded successfully!"})
        except Semester.DoesNotExist:
            return Response({"error": "Invalid Semester ID"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



        
# manually activate and inactivate school or college

from django.utils import timezone

class ToggleInstitutionActivationView(APIView):
    def post(self, request, institution_type, pk):
        print(f"Looking for institution type: {institution_type}, id: {pk}")  # Debug print
        
        try:
            if institution_type == 'school':
                institution = School.objects.get(instution_obj__id=pk)
            elif institution_type == 'college':
                institution = College.objects.get(instution_obj__id=pk)
            else:
                return Response({"detail": "Invalid institution type."}, status=status.HTTP_400_BAD_REQUEST)

            print("Found institution:", institution)

            user = institution.instution_obj.user_object
            print(f"User found: {user}")

            payment = Payment.objects.filter(user_obj=user, is_paid=True).order_by('-end_date').first()
            print(f"Latest payment: {payment}")

            if not payment or not payment.end_date or payment.end_date < timezone.now().date():
                return Response(
                    {"detail": "Cannot activate institution. Subscription is not active."},
                    status=status.HTTP_403_FORBIDDEN
                )

            institution.is_manually_deactivated = not institution.is_manually_deactivated
            institution.is_active = not institution.is_manually_deactivated
            institution.save()

            return Response({
                "id": institution.id,
                "is_active": institution.is_active,
                "is_manually_deactivated": institution.is_manually_deactivated,
            }, status=status.HTTP_200_OK)

        except (School.DoesNotExist, College.DoesNotExist):
            return Response({"detail": "Institution not found."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"Unexpected error: {e}")
            return Response({"detail": "Internal server error."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


#ck editor for self editing contents
class LandingPageContentView(generics.RetrieveUpdateAPIView):
    queryset = LandingPageContent.objects.all()
    serializer_class = LandingPageContentSerializer

    def get_object(self):
        return LandingPageContent.objects.first()
    
    
class CouponCreateListView(generics.ListCreateAPIView):
    queryset = Coupon.objects.all()
    serializer_class = CouponSerializer
    permission_classes = [IsSuperadminOrReadOnlyForStaff]
    authentication_classes=[authentication.TokenAuthentication]
    
    
class CouponDeleteView(generics.DestroyAPIView):
    queryset = Coupon.objects.all()
    serializer_class = CouponSerializer
    permission_classes = [permissions.IsAdminUser]
    authentication_classes=[authentication.TokenAuthentication]
    
    
class ApplyCouponView(APIView):
    def post(self, request):
        code = request.data.get("coupon_code")
        old_package_id = request.data.get("old_package_id")
        new_package_id = request.data.get("new_package_id")
        original_amount = request.data.get("amount")

        if not code or not old_package_id or not new_package_id or not original_amount:
            return Response({"error": "Missing required fields"}, status=status.HTTP_400_BAD_REQUEST)

        # Parse amount safely
        try:
            amount = float(original_amount)
        except (ValueError, TypeError):
            return Response({"error": "Invalid amount"}, status=status.HTTP_400_BAD_REQUEST)

        # Get coupon
        try:
            coupon = Coupon.objects.get(code=code, is_active=True)
        except Coupon.DoesNotExist:
            return Response({"error": "Invalid or inactive coupon"}, status=status.HTTP_400_BAD_REQUEST)

        # Get packages
        try:
            old_package = SubscriptionPackage.objects.get(id=old_package_id)
            new_package = SubscriptionPackage.objects.get(id=new_package_id)
        except SubscriptionPackage.DoesNotExist:
            return Response({"error": "Invalid package"}, status=status.HTTP_400_BAD_REQUEST)

        # Validate coupon
        if not coupon.is_valid_for_upgrade(old_package, new_package):
            return Response({"error": "Coupon not valid for this upgrade"}, status=status.HTTP_400_BAD_REQUEST)

        # Apply discount (reuse model logic if added)
        if coupon.discount_type == "fixed":
            final_amount = max(amount - float(coupon.discount_value), 0)
        elif coupon.discount_type == "percentage":
            final_amount = amount - (amount * float(coupon.discount_value) / 100)
        else:
            final_amount = amount

        final_amount = round(final_amount, 2)

        return Response({
            "success": True,
            "coupon": coupon.code,
            "original_amount": amount,
            "discount_value": str(coupon.discount_value),
            "discount_type": coupon.discount_type,
            "discount_applied": round(amount - final_amount, 2),
            "final_amount": final_amount,
        }, status=status.HTTP_200_OK)




class ProfileView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user

        # Case 1: Django createsuperuser
        if user.is_superuser:
            role = "superadmin"
        # Case 2: Normal users with role field
        elif hasattr(user, "role") and user.role:
            role = user.role
        else:
            role = "unknown"

        profile_data = {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "role": role,
        }
        return Response(profile_data)